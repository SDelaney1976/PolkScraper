from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
import datetime
import time
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

REQUIRED_COLS_ORDER = [
    "Name", "Address 1", "City", "State", "Zip",
    "Sex", "Race", "Public Defender", "Capture Date"
]

def ensure_required_columns(df: pd.DataFrame, set_capture_date: bool = False) -> pd.DataFrame:
    """Add any missing required columns and place them at the end in REQUIRED_COLS_ORDER."""
    # Add missing required columns
    for col in REQUIRED_COLS_ORDER:
        if col not in df.columns:
            df[col] = ""

    # Optionally set Capture Date for these rows (today, MM/DD/YYYY)
    if set_capture_date:
        today_str = datetime.date.today().strftime("%m/%d/%Y")
        df["Capture Date"] = today_str

    # Reorder so required columns are always at the end in the exact order
    non_required = [c for c in df.columns if c not in REQUIRED_COLS_ORDER]
    df = df[non_required + REQUIRED_COLS_ORDER]
    return df

def setup_driver():
    options = Options()
    options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
    return webdriver.Chrome(options=options)

def get_monday_to_today():
    today = datetime.date.today()
    monday = today - datetime.timedelta(days=today.weekday())
    return monday.strftime("%m/%d/%Y"), today.strftime("%m/%d/%Y")

def extract_case_data(driver):
    wait = WebDriverWait(driver, 25)
    wait.until(EC.presence_of_element_located(
        (By.XPATH, "//table[@id='DataTables_Table_0']//tbody//tr")
    ))

    rows = driver.find_elements(By.XPATH, "//table[@id='DataTables_Table_0']//tbody//tr")
    case_data = []
    for row in rows:
        try:
            #case_number = row.find_element(By.XPATH, "./td[4]/a").text.strip()
            case_number = row.find_element(By.XPATH, ".//a").get_attribute("title").strip()
            status = row.find_element(By.XPATH, "./td[3]").text.strip()
            case_data.append({"Case Number": case_number, "Status": status})
        except:
            continue
    return case_data

def paginate_and_collect_all_case_data(driver):
    all_data = []

    all_data.extend(extract_case_data(driver))

    pagination_buttons = driver.find_elements(By.XPATH, "//ul[@class='pagination']/li/a")
    total_pages = len(pagination_buttons)

    for page_index in range(1, total_pages):
        try:
            pagination = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located(
                (By.XPATH, "//ul[@class='pagination']/li/a")
            ))
            next_button = pagination[page_index]
            driver.execute_script("arguments[0].click();", next_button)

            time.sleep(1)
            all_data.extend(extract_case_data(driver))
        except Exception as e:
            print(f"⚠️ Failed to process page {page_index + 1}: {e}")
            continue

    return all_data

def run_case_search_and_export(from_date, to_date, court_type, court_abbr, case_number="0"):
    url = "https://showcase.polkcountyclerk.net/showcaseweb/"
    driver = setup_driver()
    wait = WebDriverWait(driver, 30)

    driver.get(url)

    wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Case Search"))).click()
    wait.until(EC.visibility_of_element_located((By.XPATH, "//input[@placeholder='From Date:']")))

    driver.find_element(By.XPATH, "//input[@placeholder='Case Number']").send_keys(case_number)
    driver.find_element(By.XPATH, "//input[@placeholder='From Date:']").send_keys(from_date)
    driver.find_element(By.XPATH, "//input[@placeholder='To Date:']").send_keys(to_date)

    court_type_dropdown = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//label[contains(text(),'Court Type')]/following-sibling::select")))
    Select(court_type_dropdown).select_by_visible_text(court_type)

    driver.find_element(By.XPATH, "//button[contains(text(),'Search')]").click()

    case_data = paginate_and_collect_all_case_data(driver)

    print(f"🔍 Total cases scraped: {len(case_data)}")
    print("📄 Sample scraped case numbers:")
    for case in case_data[:5]:
        print("-", case["Case Number"])

    driver.quit()

    if not case_data:
        print("⚠️ No case data found.")
        return

    # Output filename based on Monday-of-week + court abbr
    monday_date = datetime.date.today() - datetime.timedelta(days=datetime.date.today().weekday())
    formatted_date = monday_date.strftime("%m_%d_%Y")
    filename = f"{court_abbr}_{formatted_date}_cases.xlsx"

    # New data
    new_df = pd.DataFrame(case_data)
    # Ensure required columns for NEW rows (also sets Capture Date)
    new_df = ensure_required_columns(new_df, set_capture_date=True)

    if os.path.exists(filename):
        # Load existing file
        existing_df = pd.read_excel(filename)

        # If existing file lacks required columns, fix it and rewrite once
        need_rewrite = any(col not in existing_df.columns for col in REQUIRED_COLS_ORDER)
        if need_rewrite:
            existing_df = ensure_required_columns(existing_df, set_capture_date=False)
            # Rewrite the whole file with updated columns/order
            existing_df.to_excel(filename, index=False)

        # Deduplicate by Case Number against (possibly rewritten) existing_df
        existing_case_numbers = set(existing_df['Case Number'].astype(str))
        new_df_filtered = new_df[~new_df['Case Number'].astype(str).isin(existing_case_numbers)]

        if not new_df_filtered.empty:
            # Align columns to the existing file's column order (so headerless append is safe)
            aligned_cols = list(existing_df.columns)
            # Add any truly new columns (unlikely, but safe-guard)
            for c in new_df_filtered.columns:
                if c not in aligned_cols:
                    aligned_cols.append(c)
            new_df_filtered = new_df_filtered.reindex(columns=aligned_cols, fill_value="")

            # Append without headers
            start_row = len(existing_df) + 1
            with pd.ExcelWriter(filename, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                new_df_filtered.to_excel(writer, index=False, header=False, startrow=start_row)

            # Highlight appended rows green
            wb = load_workbook(filename)
            ws = wb.active
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            # Use number of columns from the sheet header row
            num_cols = ws.max_column
            # Appended region: rows start_row+1 ... start_row+len(new_df_filtered)
            for row in range(start_row + 1, start_row + 1 + len(new_df_filtered)):
                for col in range(1, num_cols + 1):
                    ws.cell(row=row, column=col).fill = green_fill
            wb.save(filename)

            print(f"✅ Appended {len(new_df_filtered)} new cases to '{filename}' (highlighted in green)")
        else:
            print("✅ No new cases found — file already contains all cases.")
    else:
        # First time creating the file — ensure required columns and save
        new_df = ensure_required_columns(new_df, set_capture_date=True)
        new_df.to_excel(filename, index=False)

        # Highlight all rows green (excluding header)
        wb = load_workbook(filename)
        ws = wb.active
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        num_cols = ws.max_column
        for row in range(2, len(new_df) + 2):
            for col in range(1, num_cols + 1):
                ws.cell(row=row, column=col).fill = green_fill
        wb.save(filename)

        print(f"✅ Created '{filename}' with {len(new_df)} new cases (highlighted in green)")

def main():
    print("Enter the date range for your search.")
    print("Leave blank to default to this week's Monday through today.\n")

    from_input = input("From Date (MM/DD/YYYY): ").strip()
    to_input = input("To Date (MM/DD/YYYY): ").strip()

    if not from_input or not to_input:
        from_date, to_date = get_monday_to_today()
        print(f"📅 Using default range: {from_date} to {to_date}")
    else:
        try:
            datetime.datetime.strptime(from_input, "%m/%d/%Y")
            datetime.datetime.strptime(to_input, "%m/%d/%Y")
            from_date = from_input
            to_date = to_input
        except ValueError:
            print("❌ Invalid date format. Please enter dates as MM/DD/YYYY.")
            return

    court_type_map = {
        "MM": "Misdemeanor",
        "CT": "Criminal Traffic",
        "CF": "Circuit Criminal"
    }

    court_input = input("Enter Court Type (MM = Misdemeanor, CT = Criminal Traffic, CF = Circuit Criminal): ").strip().upper()
    court_type = court_type_map.get(court_input)

    if not court_type:
        print("❌ Invalid court type. Please enter MM, CT, or CF.")
        return

    run_case_search_and_export(from_date, to_date, court_type, court_input)

if __name__ == "__main__":
    main()
