import os
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from copy import copy
from datetime import datetime

def process_case(driver, wait, case_number):
    try:
        wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(),'Case Search')]"))).click()
        time.sleep(1)

        wait.until(EC.visibility_of_element_located((By.XPATH, "//input[@aria-label='case number']")))
        input_box = driver.find_element(By.XPATH, "//input[@aria-label='case number']")
        input_box.clear()
        input_box.send_keys(case_number)
        driver.find_element(By.XPATH, "//button[contains(text(),'Search')]").click()

        # Wait for results
        wait.until(EC.presence_of_element_located((By.XPATH, "//table[contains(@id,'DataTables_Table')]//tbody/tr")))
        rows = driver.find_elements(By.XPATH, "//table[contains(@id,'DataTables_Table')]//tbody/tr")
        if not rows:
            print("⚠️ No rows found in results table.")
            return None

        row = rows[0]
        cells = row.find_elements(By.TAG_NAME, "td")

        if not cells or len(cells) < 19:
            print(f"⚠️ Row found but only {len(cells)} columns — expected at least 19.")
            return None

        data = {
            "Case Number": cells[1].text.strip(),
            "Status": cells[2].text.strip(),
            "Name": cells[5].text.strip(),
            "Address 1": driver.execute_script("return arguments[0].textContent;", cells[15]).strip(),
            "City": driver.execute_script("return arguments[0].textContent;", cells[16]).strip(),
            "State": driver.execute_script("return arguments[0].textContent;", cells[17]).strip(),
            "Zip": driver.execute_script("return arguments[0].textContent;", cells[18]).strip(),
            "Sex": cells[9].text.strip(),
            "Race": cells[8].text.strip(),
            "Public Defender": "",  # Placeholder for later
        }

        print(f"✅ Extracted data for {case_number}: {data}")
        return data

    except Exception as e:
        print(f"❌ Failed to process {case_number}: {e}")
        return None

def main():
    Tk().withdraw()
    file_path = askopenfilename(title="Select Excel file", filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        print("No file selected.")
        return

    today_str = datetime.now().strftime("%m/%d/%Y")
    df = pd.read_excel(file_path)

    for col in ["Case Number", "Status", "Name", "Address 1", "City", "State", "Zip", "Sex", "Race", "Public Defender", "Capture Date"]:
        if col not in df.columns:
            df[col] = ""

    wb = load_workbook(file_path)
    ws = wb.active

    row_fills = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row_fills.append([copy(cell.fill) for cell in row])

    options = Options()
    options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
    driver = webdriver.Chrome(options=options)
    wait = WebDriverWait(driver, 20)

    for idx, row in df.iterrows():
        case_number = row["Case Number"]
        if pd.notna(row.get("Name")) and row["Name"]:
            continue

        print(f"🔍 Processing {case_number}")
        result = process_case(driver, wait, case_number)

        if result:
            for key, val in result.items():
                df.at[idx, key] = val
            df.at[idx, "Capture Date"] = today_str

            for col_idx, key in enumerate(df.columns):
                ws.cell(row=idx + 2, column=col_idx + 1).value = df.iloc[idx][key]
                try:
                    ws.cell(row=idx + 2, column=col_idx + 1).fill = row_fills[idx][col_idx]
                except:
                    pass

            wb.save(file_path)

    print("✅ All cases processed — Chrome remains open!")

if __name__ == "__main__":
    main()
