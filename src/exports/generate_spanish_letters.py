# src/exports/generate_spanish_letters.py
import os
import sys
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import pandas as pd

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


# --- match app.py Output location in dev & frozen builds ---
def get_export_dir() -> Path:
    """
    Match app.py logic so dev & frozen builds write to <repo_root>/Output.
    """
    if getattr(sys, "frozen", False):
        # .../YourApp.app/Contents/MacOS/<exe> → go up to repo root
        repo_root = Path(sys.executable).resolve().parents[4]
    else:
        # this file is .../repo/src/exports/<file>.py → up two to reach repo root
        repo_root = Path(__file__).resolve().parents[2]
    out_dir = repo_root / "Output"
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir


def _pick_excel(title="Select source Excel (the merged or raw file)"):
    root = tk.Tk(); root.withdraw()
    path = filedialog.askopenfilename(
        title=title,
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"), ("All files", "*.*")]
    )
    root.update(); root.destroy()
    return Path(path) if path else None


def _ask_date(default_str):
    root = tk.Tk(); root.withdraw()
    s = simpledialog.askstring(
        "Capture Date",
        "Enter the capture date (YYYY-MM-DD) for the output filename:",
        initialvalue=default_str,
        parent=None
    )
    root.update(); root.destroy()
    return s.strip() if s else None


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _filter_hispanic(df: pd.DataFrame) -> pd.DataFrame:
    if "Race" not in df.columns:
        raise ValueError("The selected Excel does not contain a 'Race' column.")
    race = df["Race"].astype(str).str.strip().str.lower()
    return df[race.eq("hispanic")]


def _write_with_table(df: pd.DataFrame, out_path: Path,
                      sheet_name="MergeData", table_name="MergeDataTable"):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(out_path)
    ws = wb[sheet_name]

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 2:
        wb.save(out_path)
        return

    from openpyxl.utils import get_column_letter
    ref = f"A1:{get_column_letter(max_col)}{max_row}"

    # ensure unique table name
    tn = table_name
    existing = {t.displayName for t in ws._tables}
    i = 2
    while tn in existing:
        tn = f"{table_name}_{i}"
        i += 1

    table = Table(displayName=tn, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(out_path)


def _open_word_template_picker():
    root = tk.Tk(); root.withdraw()
    path = filedialog.askopenfilename(
        title="Open the Spanish Word letter template (.docx/.docm)",
        filetypes=[("Word documents", "*.docx *.docm"), ("All files", "*.*")]
    )
    root.update(); root.destroy()
    return Path(path) if path else None


def main():
    # 1) Pick source Excel
    src = _pick_excel("Select the Excel to pull Hispanic rows from")
    if not src:
        print("Canceled: no Excel selected.")
        return

    # 2) Ask for filename date
    from datetime import date
    default_date = date.today().strftime("%Y-%m-%d")
    cap_date = _ask_date(default_date)
    if not cap_date:
        print("Canceled: no capture date provided.")
        return

    # 3) Read → normalize → filter
    try:
        df = pd.read_excel(src, dtype=str)
        df = _normalize_columns(df)
        df_out = _filter_hispanic(df)
    except Exception as e:
        messagebox.showerror("Read/Filter Error", f"Could not process Excel:\n{e}")
        return

    if df_out.empty:
        messagebox.showinfo("No Rows", "No Hispanic rows found in the selected file.")
        return

    # 4) Save to <repo>/Output/Spanish_MergeData_<date>.xlsx
    OUTPUT_DIR = get_export_dir()
    out_path = OUTPUT_DIR / f"Spanish_MergeData_{cap_date}.xlsx"

    try:
        _write_with_table(df_out, out_path, sheet_name="MergeData", table_name="MergeDataTable")
    except Exception as e:
        messagebox.showerror("Write Error", f"Failed writing output Excel:\n{e}")
        return

    # 5) Prompt to open the Word template
    messagebox.showinfo("Export complete", f"Saved:\n{out_path}\n\nNext: pick your Spanish Word template.")
    tpl = _open_word_template_picker()
    if tpl:
        try:
            if sys.platform == "darwin":
                os.system(f"open '{tpl.as_posix()}'")
            elif sys.platform.startswith("win"):
                os.startfile(str(tpl))  # type: ignore[attr-defined]
            else:
                os.system(f"xdg-open '{tpl.as_posix()}'")
        except Exception:
            pass
    else:
        print("Template open canceled.")


if __name__ == "__main__":
    main()
