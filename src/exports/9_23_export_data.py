# src/exports/export_data.py
from __future__ import annotations

import sys
from pathlib import Path
from datetime import date
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# GUI bits for picking files / prompting for date
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog


# -------------------------------
# Output location helper (DEV + APP)
# -------------------------------
def _output_dir() -> Path:
    """
    Dev:  <repo>/Output
    App:  ~/Documents/PolkScraper/Output
    """
    if getattr(sys, "frozen", False):
        base = Path.home() / "Documents" / "PolkScraper" / "Output"
    else:
        # this file is .../repo/src/exports/export_data.py → go up two to repo
        base = Path(__file__).resolve().parents[2] / "Output"
    base.mkdir(parents=True, exist_ok=True)
    return base


# -------------------------------
# Small GUI helpers
# -------------------------------
def _pick_sources() -> List[Path]:
    root = tk.Tk(); root.withdraw()
    paths = filedialog.askopenfilenames(
        title="Select 1–3 Excel files to merge for mail-merge",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"), ("All files", "*.*")]
    )
    root.update(); root.destroy()
    return [Path(p) for p in paths or []]


def _ask_date(default_str: str) -> Optional[str]:
    root = tk.Tk(); root.withdraw()
    s = simpledialog.askstring(
        "Capture Date",
        "Enter the capture date (YYYY-MM-DD) for the output filename:",
        initialvalue=default_str,
        parent=None
    )
    root.update(); root.destroy()
    return s.strip() if s else None


# -------------------------------
# Data munging
# -------------------------------
_EXPECTED_ORDER = [
    "Name", "Address 1", "Address 2", "City", "State", "Zip",
    "Case Number", "Status", "Sex", "Race", "Phone", "Public Defender",
    "Capture Date",
]

def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).strip() for c in df.columns]
    return df

def _read_one(path: Path) -> pd.DataFrame:
    # Keep dtype=str to avoid pandas guessing types inconsistently across files.
    df = pd.read_excel(path, dtype=str)
    return _normalize_columns(df)

def _concat_frames(paths: List[Path]) -> pd.DataFrame:
    parts = []
    for p in paths:
        try:
            parts.append(_read_one(p))
        except Exception as e:
            messagebox.showerror("Read Error", f"Could not read {p}:\n{e}")
            raise
    if not parts:
        return pd.DataFrame(columns=_EXPECTED_ORDER)
    df = pd.concat(parts, ignore_index=True)
    df = _normalize_columns(df)

    # Re-order columns where possible; keep any extras at the end
    existing = [c for c in _EXPECTED_ORDER if c in df.columns]
    remaining = [c for c in df.columns if c not in existing]
    df = df[existing + remaining]

    return df


# -------------------------------
# Capture Date normalizer (robust)
# -------------------------------
def _normalize_capture_date(df: pd.DataFrame, col: str = "Capture Date") -> pd.Series:
    """
    Return a Series of pure date objects parsed from df[col].
    Handles strings (MM/DD/YYYY, YYYY-MM-DD), datetimes, and Excel serials.
    """
    if col not in df.columns:
        raise KeyError("Missing 'Capture Date' column in the input file(s).")

    s = df[col]

    # If already datetime-like (rare since we read as str), handle quickly
    if pd.api.types.is_datetime64_any_dtype(s):
        return s.dt.date

    # Numeric → Excel serials
    if pd.api.types.is_numeric_dtype(s):
        excel = pd.to_datetime("1899-12-30") + pd.to_timedelta(s, unit="D")
        return excel.dt.date

    # Strings / mixed → robust parse
    parsed = pd.to_datetime(s, errors="coerce", infer_datetime_format=True)

    # Fallback for cells that look numeric but arrived as strings
    mask_numlike = parsed.isna() & s.astype(str).str.fullmatch(r"\d+(\.\d+)?", na=False)
    if mask_numlike.any():
        numeric = pd.to_numeric(s.where(mask_numlike), errors="coerce")
        excel = pd.to_datetime("1899-12-30") + pd.to_timedelta(numeric, unit="D")
        parsed = parsed.fillna(excel)

    return parsed.dt.date


# -------------------------------
# Excel writer w/ Table for Word
# -------------------------------
def _write_with_table(df: pd.DataFrame, out_path: Path,
                      sheet_name="MergeData", table_name="MergeDataTable") -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Write with pandas (openpyxl engine auto)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Add an Excel Table so Word mail-merge can pick by name
    wb = load_workbook(out_path)
    ws = wb[sheet_name]

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 2:
        wb.save(out_path)
        return

    from openpyxl.utils import get_column_letter
    end_col = get_column_letter(max_col)
    ref = f"A1:{end_col}{max_row}"

    # Ensure unique table name if file gets overwritten
    existing = {t.displayName for t in ws._tables}
    tn = table_name
    i = 2
    while tn in existing:
        tn = f"{table_name}_{i}"
        i += 1

    table = Table(displayName=tn, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(out_path)


# -------------------------------
# Public entrypoint used by app.py
# -------------------------------
def export_today_merge_data_with_gui() -> Path:
    """
    Workflow:
      1) User picks 1–3 Excel files (any order).
      2) Prompt for date (default = today).
      3) Merge, filter rows to that Capture Date, and save to
         PolkScraper/Output/MergeData_<date>.xlsx
         (sheet 'MergeData', table 'MergeDataTable').
    Returns the Path to the written file.
    """
    # 1) Pick sources
    sources = _pick_sources()
    if not sources:
        raise RuntimeError("No source files selected.")

    if len(sources) > 3:
        messagebox.showwarning("Too many files", "Please select no more than 3 files.")
        sources = sources[:3]

    # 2) Ask for capture date (YYYY-MM-DD)
    default_date = date.today().strftime("%Y-%m-%d")
    cap_date = _ask_date(default_date)
    if not cap_date:
        raise RuntimeError("Export canceled: no capture date entered.")

    # Validate/convert the entered date once
    try:
        target_date = pd.to_datetime(cap_date, errors="raise").date()
    except Exception:
        raise RuntimeError(f"Invalid capture date: {cap_date!r}. Use YYYY-MM-DD.")

    # 3) Merge
    df = _concat_frames(sources)
    if df.empty:
        raise RuntimeError("No rows to export after merging.")

    # 4) Filter by Capture Date == target_date
    norm_dates = _normalize_capture_date(df, "Capture Date")
    mask = norm_dates == target_date
    filtered = df.loc[mask].copy()

    if filtered.empty:
        messagebox.showwarning(
            "No rows for selected date",
            f"No rows found where Capture Date == {target_date:%m/%d/%Y}.\n"
            "Nothing will be exported."
        )
        raise RuntimeError("Export canceled: no rows matched the selected Capture Date.")

    # 5) Write output
    out_dir = _output_dir()
    out_path = out_dir / f"MergeData_{cap_date}.xlsx"

    _write_with_table(filtered, out_path, sheet_name="MergeData", table_name="MergeDataTable")

    # On macOS, optionally reveal in Finder (safe no-op elsewhere if desired by caller)
    return out_path


# -------------------------------
# CLI test hook (optional)
# -------------------------------
if __name__ == "__main__":
    try:
        path = export_today_merge_data_with_gui()
        print(f"✅ Exported: {path}")
        # macOS: reveal
        try:
            import subprocess
            if sys.platform == "darwin":
                subprocess.run(["open", "-R", str(path)], check=False)
        except Exception:
            pass
    except Exception as e:
        print(f"❌ Export failed: {e}")
