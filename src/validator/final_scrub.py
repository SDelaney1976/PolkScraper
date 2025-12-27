# src/validator/final_scrub.py
import sys
import os
import re
import datetime
from tkinter import Tk, filedialog

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from dateutil.parser import parse as date_parse
from typing import List, Tuple, Dict

# --- Name formatting helpers (drop-in) ---------------------------------------
# Keep your SUFFIXES list if you already defined it; otherwise:
SUFFIXES_STD = {"JR": "Jr.", "SR": "Sr."}
SUFFIXES_ROMAN = {"II", "III", "IV", "V"}  # keep uppercase, no period

# Particles typically lowercased in many styles (adjust to taste)
LOWER_PARTICLES = {
    "da", "de", "del", "della", "di", "du", "la", "le", "lo",
    "van", "von", "bin", "ibn", "al", "dos", "das", "de la", "de los"
}

def _clean_spaces(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip())

def _std_suffix(token: str) -> str:
    t = re.sub(r"\.", "", token or "").upper()
    if t in SUFFIXES_STD:
        return SUFFIXES_STD[t]
    if t in SUFFIXES_ROMAN:
        return t  # e.g., III
    return token  # unknown suffix returned as-is

def _is_suffix(token: str) -> bool:
    if not token:
        return False
    t = re.sub(r"\.", "", token).upper()
    return (t in SUFFIXES_STD) or (t in SUFFIXES_ROMAN)

def _proper_case_token_word(w: str) -> str:
    """
    Proper-case a single word, handling:
      - O'/D'/L' (e.g., O'BRIAN -> O'Brian)
      - McX... (e.g., McDONALD -> McDonald)
      - Initials and acronyms
    """
    if not w:
        return w

    # If it's an initial like "J" or "J." → "J."
    if re.fullmatch(r"[A-Za-z]\.?", w):
        return w[0].upper() + "."

    # O'/D'/L'
    m = re.match(r"^([A-Za-z])'([A-Za-z].*)$", w)
    if m:
        head = m.group(1).upper()
        tail = m.group(2).lower().title()
        # Special-case “O” and “D” and “L”
        if head in {"O", "D", "L"}:
            # Keep apostrophe, title-case remainder
            return f"{head}'{tail}"
        return f"{head}'{tail}"

    # McX... → McX...
    m2 = re.match(r"^mc([A-Za-z])(.*)$", w, flags=re.IGNORECASE)
    if m2:
        return "Mc" + m2.group(1).upper() + m2.group(2).lower()

    # Preserve all-caps acronyms ≥ 2 letters if they look like acronyms
    if len(w) >= 2 and w.isupper():
        if re.fullmatch(r"[A-Z][A-Z\-']*", w):
            return w[0].upper() + w[1:].lower()
        return w

    return w[0].upper() + w[1:].lower()

def _fix_mac_token(token: str) -> str:
    """
    If a (hyphen segment of a) last-name token starts with 'Mac' (any case),
    force 'Mac' + next letter uppercase + rest lowercase.
    Examples:
      Macdonald -> MacDonald
      MACARTHUR -> MacArthur
      Macdonald-SMITH -> MacDonald-Smith
    """
    if not token:
        return token

    parts = token.split("-")
    out_parts = []
    for p in parts:
        m = re.match(r"^mac([A-Za-z])(.*)$", p, flags=re.IGNORECASE)
        if m:
            out_parts.append("Mac" + m.group(1).upper() + m.group(2).lower())
        else:
            out_parts.append(p)
    return "-".join(out_parts)


def _proper_case_token(token: str) -> str:
    """
    Proper-case a hyphenated token and apply particle rules later.
    """
    if not token:
        return token
    parts = token.split("-")
    parts = [_proper_case_token_word(p) for p in parts]
    return "-".join(parts)

def _lowercase_particles_when_appropriate(tokens: List[str]) -> List[str]:
    """
    Lowercases particles when they appear as separate tokens in the last name.
    Handles multi-word particles like 'de la'.
    """
    if not tokens:
        return tokens

    out = tokens[:]
    i = 0
    while i < len(out):
        # Handle two-word particle (e.g., "de la")
        if i + 1 < len(out):
            pair = f"{out[i].lower()} {out[i+1].lower()}"
            if pair in LOWER_PARTICLES:
                out[i] = out[i].lower()
                out[i+1] = out[i+1].lower()
                i += 2
                continue
        # Single-word particle
        if out[i].lower() in LOWER_PARTICLES:
            out[i] = out[i].lower()
        i += 1
    return out

def _split_commas(name: str) -> Tuple[str, str]:
    """
    If there's a comma, assume 'Last, First ...'.
    Return (last_part, first_part_plus).
    """
    if "," in name:
        left, right = name.split(",", 1)
        return _clean_spaces(left), _clean_spaces(right)
    return "", name

def _extract_suffix(tokens: List[str]) -> Tuple[List[str], str]:
    """
    Pop a recognized suffix from the end of tokens.
    """
    if not tokens:
        return tokens, ""
    last = tokens[-1]
    if _is_suffix(last):
        return tokens[:-1], _std_suffix(last)
    return tokens, ""

def _proper_case_name_tokens(tokens: List[str], is_last_name: bool = False) -> List[str]:
    if not tokens:
        return tokens
    cased = [_proper_case_token(t) for t in tokens]
    if is_last_name:
        # Apply Mac rule only on surnames to avoid false positives on first names
        cased = [_fix_mac_token(t) for t in cased]
        cased = _lowercase_particles_when_appropriate(cased)
    return cased

def _join_nonempty(parts: List[str]) -> str:
    return " ".join([p for p in parts if p])

def parse_name_parts(raw: str) -> Dict[str, str]:
    """
    Parses a name string that might be in either:
      - 'LAST, FIRST M [SUFFIX]'
      - 'FIRST M LAST [SUFFIX]'
    Returns dict: first, middle, last, suffix, full
    """
    if not isinstance(raw, str):
        return {"first": "", "middle": "", "last": "", "suffix": "", "full": ""}

    s = _clean_spaces(raw)
    if not s:
        return {"first": "", "middle": "", "last": "", "suffix": "", "full": ""}

    # Try comma form first
    last_side, right = _split_commas(s)

    if last_side:
        last_tokens = _clean_spaces(last_side).split(" ")
        right_tokens = _clean_spaces(right).split(" ")
        right_tokens, suffix = _extract_suffix(right_tokens)

        # First + (optional) middle
        first = right_tokens[0] if right_tokens else ""
        middle = _join_nonempty(right_tokens[1:]) if len(right_tokens) > 1 else ""

        last_tokens = _proper_case_name_tokens(last_tokens, is_last_name=True)
        first_tokens = _proper_case_name_tokens([first])
        middle_tokens = _proper_case_name_tokens(middle.split(" ")) if middle else []

        first_out = _join_nonempty(first_tokens)
        middle_out = _join_nonempty(middle_tokens)
        last_out = _join_nonempty(last_tokens)

        full = _join_nonempty([first_out, middle_out, last_out, suffix])
        return {
            "first": first_out,
            "middle": middle_out,
            "last": last_out,
            "suffix": suffix,
            "full": full
        }

    # No comma → assume First [Middle ...] Last [Suffix]
    tokens = s.split(" ")
    tokens, suffix = _extract_suffix(tokens)
    if len(tokens) == 1:
        # Single token; treat as last name
        last_tokens = _proper_case_name_tokens(tokens, is_last_name=True)
        full = _join_nonempty([_join_nonempty(last_tokens), suffix])
        return {"first": "", "middle": "", "last": last_tokens[0], "suffix": suffix, "full": full}

    # First token is first name; last token is last name; middle = rest
    first = tokens[0]
    last = tokens[-1]
    middle = tokens[1:-1]

    first_tokens = _proper_case_name_tokens([first])
    last_tokens = _proper_case_name_tokens([last], is_last_name=True)
    middle_tokens = _proper_case_name_tokens(middle) if middle else []

    first_out = _join_nonempty(first_tokens)
    middle_out = _join_nonempty(middle_tokens)
    last_out = _join_nonempty(last_tokens)

    full = _join_nonempty([first_out, middle_out, last_out, suffix])
    return {
        "first": first_out,
        "middle": middle_out,
        "last": last_out,
        "suffix": suffix,
        "full": full
    }

# --- NEW: targeted “O” prefix apostrophe fixer + post-processing -------------

# High-confidence Irish stems where O' is commonly correct (case-insensitive)
_O_STEMS = {
    "neal": "Neal", "neil": "Neil", "neill": "Neill",
    "brien": "Brien", "brian": "Brian",
    "connor": "Connor", "conner": "Conner",
    "donnell": "Donnell", "dwyer": "Dwyer",
    "hara": "Hara", "hare": "Hare",
    "hagan": "Hagan",
    "keefe": "Keefe",
    "leary": "Leary",
    "reilly": "Reilly", "riley": "Riley", "regan": "Regan",
    "rourke": "Rourke",
    "shea": "Shea", "shaughnessy": "Shaughnessy",
    "sullivan": "Sullivan",
    "toole": "Toole",
    "gorman": "Gorman", "gara": "Gara",
    "malley": "Malley",
}

_SUFFIX_TOKENS = {"Jr.", "Sr.", "II", "III", "IV", "V"}

def _fix_missing_apostrophe_o_prefix(token: str) -> str:
    """If token looks like Oneal/Obrien/Oconnor/etc (no apostrophe),
    convert to O'Neal/O'Brien/O'Connor, using a safe whitelist of stems."""
    if "'" in token:
        return token
    low = token.lower()
    m = re.match(r"^o([a-z]+)$", low)
    if not m:
        return token
    stem = m.group(1)
    if stem in _O_STEMS:
        return "O'" + _O_STEMS[stem]
    return token

def _postprocess_name(final_name: str) -> str:
    """
    Apply three tweaks to the final formatted name:
      1) Remove period after single-letter middle initials (Sean P. Delaney -> Sean P Delaney)
      2) Remove hyphens in names (Juan-Carlos -> Juan Carlos)
      3) Insert missing apostrophes for certain O-prefixed surnames (Oneal -> O'Neal, etc.)
    """
    if not final_name:
        return final_name

    # 1) middle-initial periods → remove (but keep suffix periods like Jr.)
    # Only remove when it's a standalone initial (e.g., ' P. ')
    name = re.sub(r"\b([A-Z])\.(?=\s|$)", r"\1", final_name)
    

    # 2) remove hyphens in names (turn hyphens into spaces)
    name = name.replace("-", " ")

    # normalize spaces after hyphen removal
    name = re.sub(r"\s{2,}", " ", name).strip()

    # 3) O' fixer on last-name portion (keep suffix tokens at end)
    tokens = name.split()
    if not tokens:
        return name

    # peel off suffix if present
    suffix = ""
    if tokens and tokens[-1] in _SUFFIX_TOKENS:
        suffix = " " + tokens.pop()

    if not tokens:
        return (name + suffix).strip()

    # Heuristic: apply O' fixer to the last token of the base name
    tokens[-1] = _fix_missing_apostrophe_o_prefix(tokens[-1])

    # Also apply to any internal multi-word last names (e.g., "van der Oneil")
    # This is conservative: only adjust tokens that *start* with 'O' and have no apostrophe.
    for i in range(max(1, len(tokens) - 3), len(tokens)):  # look back a bit near the end
        t = tokens[i]
        if "'" not in t and re.match(r"^O[A-Za-z]{2,}$", t):
            tokens[i] = _fix_missing_apostrophe_o_prefix(t)

    out = " ".join(tokens) + suffix
    return re.sub(r"\s{2,}", " ", out).strip()

def format_name(raw: str) -> str:
    """
    Public convenience: return proper-case 'First Middle Last Suffix'.
    """
    return parse_name_parts(raw).get("full", "")
# --- End name formatting helpers ---------------------------------------------


def standardize_race(value):
    if pd.isna(value):
        return value
    value = str(value).strip().upper()
    mapping = {
        'W': 'White', 'WHITE': 'White',
        'B': 'Black', 'BLACK': 'Black',
        'H': 'Hispanic', 'HISPANIC': 'Hispanic',
        'O': 'Other', 'OTHER': 'Other'
    }
    return mapping.get(value, value.capitalize())


def transform_address(value):
    if pd.isna(value):
        return value
    val = str(value).strip()
    val = re.sub(r'\bHighway\b', 'Hwy', val, flags=re.IGNORECASE)
    val = re.sub(r'\bBoulevard\b', 'Blvd', val, flags=re.IGNORECASE)
    val = re.sub(r'\bNortheast\b', 'NE', val, flags=re.IGNORECASE)
    val = re.sub(r'\bSoutheast\b', 'SE', val, flags=re.IGNORECASE)
    val = re.sub(r'\bNorthwest\b', 'NW', val, flags=re.IGNORECASE)
    val = re.sub(r'\bSouthwest\b', 'SW', val, flags=re.IGNORECASE)
    return val


def clean_excel_file(file_path):
    """
    Cleans a single Excel file in-place.
    Returns a dict of stats and prints human-friendly lines.
    """
    stats = {
        "file": file_path,
        "start_rows": 0,
        "removed_disposed": 0,
        "removed_814_n_kentucky": 0,
        "removed_180_e_central": 0,
        "removed_general_delivery": 0,
        "duplicates_removed": 0,
        "rows_highlighted_today": 0,
        "end_rows": 0,
        "ok": False,
        "error": None,
    }

    try:
        df = pd.read_excel(file_path)
        stats["start_rows"] = len(df)

        # Clean Name (use the robust formatter)
        if 'Name' in df.columns:
            df['Name'] = df['Name'].apply(lambda x: _postprocess_name(format_name(x)))
            print("🧼 Cleaned 'Name' column")

        # Standardize Race
        if 'Race' in df.columns:
            df['Race'] = df['Race'].apply(standardize_race)
            print("🧬 Standardized 'Race' column")

        # Remove Status == Disposed
        if 'Status' in df.columns:
            before = len(df)
            df = df[~df['Status'].astype(str).str.strip().str.lower().eq('disposed')]
            stats["removed_disposed"] = before - len(df)
            if stats["removed_disposed"]:
                print(f"🗑️ Removed {stats['removed_disposed']} row(s) with Status 'Disposed'")

        # Fix mis-typed city names
        if "City" in df.columns:
            df["City"] = df["City"].replace({"CHAMPIONS GT": "Champions Gate"})

        # Clean Address 1
        if 'Address 1' in df.columns:
            address_col = df['Address 1'].astype(str).str.strip().str.lower()
            blocked_addresses = [
                r'^814 north kentucky avenue$',
                r'^814 kentucky avenue$',
                r'^814 n kentucky avenue$',
                r'^814 n kentucky ave$',
                r'^814 kentucky ave$'
            ]
            pattern = '|'.join(blocked_addresses)
            before = len(df)
            df = df[~address_col.str.match(pattern)]
            stats["removed_814_n_kentucky"] = before - len(df)
            if stats["removed_814_n_kentucky"]:
                print(f"🏠 Removed {stats['removed_814_n_kentucky']} row(s) with '814 North Kentucky Avenue' variations")

            before = len(df)
            df = df[~df['Address 1'].astype(str).str.strip().str.lower().eq("180 east central avenue")]
            stats["removed_180_e_central"] = before - len(df)
            if stats["removed_180_e_central"]:
                print(f"🏢 Removed {stats['removed_180_e_central']} row(s) with '180 East Central Avenue'")

            before = len(df)
            df = df[~df['Address 1'].astype(str).str.lower().str.contains("general delivery")]
            stats["removed_general_delivery"] = before - len(df)
            if stats["removed_general_delivery"]:
                print(f"📦 Removed {stats['removed_general_delivery']} row(s) with 'General Delivery'")

            # Standardize keywords
            df['Address 1'] = df['Address 1'].apply(transform_address)
            print("📫 Standardized keywords in 'Address 1'")

        # Remove duplicate addresses (keep oldest Capture Date)
        if 'Address 1' in df.columns and 'Capture Date' in df.columns:
            df['Capture Date'] = pd.to_datetime(df['Capture Date'], errors='coerce').dt.date
            before = len(df)
            df = df.sort_values(by=['Address 1', 'Capture Date'])
            df = df.drop_duplicates(subset=['Address 1'], keep='first')
            stats["duplicates_removed"] = before - len(df)
            if stats["duplicates_removed"]:
                print(f"🔁 Removed {stats['duplicates_removed']} duplicate address row(s), kept oldest by Capture Date")

        # Reorder columns
        desired_order = [
            "Name", "Address 1", "Address 2", "City", "State", "Zip",
            "Case Number", "Status", "Sex", "Race", "Phone", "Public Defender"
        ]
        existing = [c for c in desired_order if c in df.columns]
        remaining = [c for c in df.columns if c not in existing]
        df = df[existing + remaining]
        print("📑 Reordered columns")

        # Final sort by Capture Date (oldest to newest)
        if 'Capture Date' in df.columns:
            df['Capture Date'] = pd.to_datetime(df['Capture Date'], errors='coerce')
            df = df.sort_values(by='Capture Date', ascending=True)
            print("📅 Sorted rows by 'Capture Date' (oldest to newest)")

        # Save cleaned data (in-place)
        df.to_excel(file_path, index=False)
        stats["end_rows"] = len(df)
        print("💾 Saved cleaned data to Excel")

        # Open workbook for formatting
        wb = load_workbook(file_path)
        ws = wb.active

        # Format Zip as text
        zip_col_index = None
        for idx, cell in enumerate(ws[1], start=1):
            if str(cell.value).strip().lower() == 'zip':
                zip_col_index = idx
                break
        if zip_col_index:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if len(row) >= zip_col_index:
                    row[zip_col_index - 1].number_format = "@"
            print("🏷️ Formatted 'Zip' column as text")

        # Highlight Capture Date == today
        date_col_index = None
        for idx, cell in enumerate(ws[1], start=1):
            if str(cell.value).strip().lower() == 'capture date':
                date_col_index = idx
                break

        rows_highlighted = 0
        green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        today = datetime.date.today()

        if date_col_index:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if len(row) < date_col_index:
                    continue
                cell = row[date_col_index - 1]
                val = cell.value
                try:
                    if isinstance(val, datetime.datetime):
                        date_val = val.date()
                    elif isinstance(val, datetime.date):
                        date_val = val
                    elif isinstance(val, str):
                        date_val = date_parse(val).date()
                    else:
                        continue
                    if date_val == today:
                        for c in row:
                            c.fill = green
                        rows_highlighted += 1
                except Exception:
                    continue
            stats["rows_highlighted_today"] = rows_highlighted
            print(f"✅ Highlighted {rows_highlighted} row(s) with today's Capture Date")
        else:
            print("⚠️ 'Capture Date' column not found — skipping highlighting")

        wb.save(file_path)
        print(f"🎉 File cleaned and saved: {file_path}")

        stats["ok"] = True
        return stats

    except Exception as e:
        stats["error"] = str(e)
        print(f"❌ Error processing file: {e}")
        return stats


def _print_banner(title):
    print("\n" + "=" * 72)
    print(title)
    print("=" * 72)


def main(argv=None):
    argv = argv if argv is not None else sys.argv[1:]

    # Get files: CLI args or picker
    if argv:
        file_paths = [p for p in argv if os.path.isfile(p)]
        missing = [p for p in argv if not os.path.isfile(p)]
        if missing:
            _print_banner("⚠️  Skipping missing paths")
            for m in missing:
                print(m)
            print()
    else:
        Tk().withdraw()
        file_paths = filedialog.askopenfilenames(
            title="Select Excel Files",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )

    if not file_paths:
        print("❌ No files selected.")
        return 1

    all_stats = []
    for fp in file_paths:
        _print_banner(f"📄 Processing: {fp}")
        stats = clean_excel_file(fp)
        # Echo a compact summary per file
        if stats["ok"]:
            print("-- Summary --")
            print(f"Start rows: {stats['start_rows']}")
            print(f"Removed Disposed: {stats['removed_disposed']}")
            print(f"Removed '814 North Kentucky Avenue': {stats['removed_814_n_kentucky']}")
            print(f"Removed '180 East Central Avenue': {stats['removed_180_e_central']}")
            print(f"Removed 'General Delivery': {stats['removed_general_delivery']}")
            print(f"Duplicate addresses removed: {stats['duplicates_removed']}")
            print(f"Rows highlighted today: {stats['rows_highlighted_today']}")
            print(f"End rows: {stats['end_rows']}")
        else:
            print("-- Summary --")
            print("FAILED:", stats.get("error", "Unknown error"))
        all_stats.append(stats)

    # Optional overall summary
    _print_banner("📊 Overall Summary")
    ok = sum(1 for s in all_stats if s["ok"])
    fail = len(all_stats) - ok
    print(f"Files processed: {len(all_stats)} (ok: {ok}, failed: {fail})")
    return 0 if fail == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
