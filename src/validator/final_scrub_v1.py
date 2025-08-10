import pandas as pd
import os
from tkinter import Tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from dateutil.parser import parse as date_parse
import datetime
import re

# Suffixes to preserve in proper-case names
SUFFIXES = ['Jr.', 'Sr.', 'II', 'III', 'IV', 'V']

def proper_case_name(name):
    if pd.isna(name):
        return name

    name = name.strip()

    # Handle names like "Doe, John Jr." → "John Doe Jr."
    if ',' in name:
        parts = name.split(',')
        if len(parts) >= 2:
            last = parts[0].strip()
            rest = parts[1].strip()
            name = f"{rest} {last}"

    # Now apply proper case and suffix logic
    tokens = name.split()
    result = []

    for token in tokens:
        if token in SUFFIXES:
            result.append(token)
        elif '-' in token:
            result.append('-'.join([t.capitalize() for t in token.split('-')]))
        else:
            result.append(token.capitalize())

    return ' '.join(result)


def standardize_race(value):
    if pd.isna(value):
        return value
    value = str(value).strip().upper()
    mapping = {
        'W': 'White',
        'WHITE': 'White',
        'B': 'Black',
        'BLACK': 'Black',
        'H': 'Hispanic',
        'HISPANIC': 'Hispanic',
        'O': 'Other',
        'OTHER': 'Other'
    }
    return mapping.get(value, value.capitalize())

def transform_address(value):
    if pd.isna(value):
        return value
    val = str(value).strip()
    val = re.sub(r'\bHighway\b', 'Hwy', val, flags=re.IGNORECASE)
    val = re.sub(r'\bBoulevard\b', 'Blvd', val, flags=re.IGNORECASE)
    val = re.sub(r'\bNortheast\b', 'NE', val, flags=re.IGNORECASE)
    val = re.sub(r'\bSoutheast\b', 'SE', val, flags=re.IGNORECASE)
    val = re.sub(r'\bNorthwest\b', 'NW', val, flags=re.IGNORECASE)
    val = re.sub(r'\bSouthwest\b', 'SW', val, flags=re.IGNORECASE)
    return val

def clean_excel_file(file_path):
    try:
        df = pd.read_excel(file_path)

        # Clean Name
        if 'Name' in df.columns:
            df['Name'] = df['Name'].apply(proper_case_name)
            print("🧼 Cleaned 'Name' column")

        # Standardize Race
        if 'Race' in df.columns:
            df['Race'] = df['Race'].apply(standardize_race)
            print("🧬 Standardized 'Race' column")

        # Remove Status == Disposed
        if 'Status' in df.columns:
            original = len(df)
            df = df[~df['Status'].astype(str).str.strip().str.lower().eq('disposed')]
            print(f"🗑️ Removed {original - len(df)} row(s) with Status 'Disposed'")

        # Clean Address 1
        if 'Address 1' in df.columns:
            original = len(df)
            df = df[~df['Address 1'].astype(str).str.strip().str.lower().eq("814 north kentucky avenue")]
            print(f"🏠 Removed {original - len(df)} row(s) with address '814 North Kentucky Avenue'")

            original = len(df)
            df = df[~df['Address 1'].astype(str).str.strip().str.lower().eq("180 east central avenue")]
            print(f"🏢 Removed {original - len(df)} row(s) with address '180 East Central Avenue'")

            original = len(df)
            df = df[~df['Address 1'].astype(str).str.lower().str.contains("general delivery")]
            print(f"📦 Removed {original - len(df)} row(s) with 'General Delivery'")

            df['Address 1'] = df['Address 1'].apply(transform_address)
            print("📫 Standardized keywords in 'Address 1'")

        # Remove duplicate addresses (keep oldest Capture Date)
        if 'Address 1' in df.columns and 'Capture Date' in df.columns:
            df['Capture Date'] = pd.to_datetime(df['Capture Date'], errors='coerce').dt.date
            before = len(df)
            df = df.sort_values(by=['Address 1', 'Capture Date'])
            df = df.drop_duplicates(subset=['Address 1'], keep='first')
            print(f"🔁 Removed {before - len(df)} duplicate address row(s), kept oldest by Capture Date")

        # Reorder columns
        desired_order = [
            "Name", "Address 1", "Address 2", "City", "State", "Zip",
            "Case Number", "Status", "Sex", "Race", "Phone", "Public Defender"
        ]
        existing = [col for col in desired_order if col in df.columns]
        remaining = [col for col in df.columns if col not in existing]
        df = df[existing + remaining]
        print("📑 Reordered columns")

        # Final sort by Capture Date (oldest to newest)
        if 'Capture Date' in df.columns:
            df['Capture Date'] = pd.to_datetime(df['Capture Date'], errors='coerce')
            df = df.sort_values(by='Capture Date', ascending=True)
            print("📅 Sorted rows by 'Capture Date' (oldest to newest)")


        # Save cleaned data
        df.to_excel(file_path, index=False)
        print("💾 Saved cleaned data to Excel")

        # Open workbook for formatting
        wb = load_workbook(file_path)
        ws = wb.active

        # Format Zip as text
        zip_col_index = None
        for idx, cell in enumerate(ws[1], start=1):
            if str(cell.value).strip().lower() == 'zip':
                zip_col_index = idx
                break
        if zip_col_index:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if len(row) >= zip_col_index:
                    row[zip_col_index - 1].number_format = "@"
            print("🏷️ Formatted 'Zip' column as text")

        # Highlight Capture Date == today
        date_col_index = None
        for idx, cell in enumerate(ws[1], start=1):
            if str(cell.value).strip().lower() == 'capture date':
                date_col_index = idx
                break

        rows_highlighted = 0
        green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        today = datetime.date.today()

        if date_col_index:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if len(row) < date_col_index:
                    continue
                cell = row[date_col_index - 1]
                val = cell.value
                try:
                    if isinstance(val, datetime.datetime):
                        date_val = val.date()
                    elif isinstance(val, datetime.date):
                        date_val = val
                    elif isinstance(val, str):
                        date_val = date_parse(val).date()
                    else:
                        continue
                    if date_val == today:
                        for c in row:
                            c.fill = green
                        rows_highlighted += 1
                except Exception:
                    continue
            print(f"✅ Highlighted {rows_highlighted} row(s) with today's Capture Date")
        else:
            print("⚠️ 'Capture Date' column not found — skipping highlighting")

        wb.save(file_path)
        print(f"🎉 File cleaned and saved: {file_path}")

    except Exception as e:
        print(f"❌ Error processing file: {e}")

if __name__ == "__main__":
    Tk().withdraw()
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if file_path:
        clean_excel_file(file_path)
    else:
        print("❌ No file selected.")
