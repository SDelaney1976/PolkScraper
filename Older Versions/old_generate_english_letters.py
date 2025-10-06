#!/usr/bin/env python3
"""
generate_english_letters.py
---------------------------
Pick a merge source Excel, filter records where Race != 'Hispanic',
save a filtered Excel into <project>/Output, then pick and open a Word template.

Workflow after this script runs:
- Word opens your template.
- In Word: Mailings → Select Recipients → Use an Existing List… → select the filtered file it just created in Output.
- Finish & Merge → Print Documents (or Edit Individual Documents to review).
"""

import sys
import datetime as dt
from pathlib import Path
from typing import Optional

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox

# Output location: same convention as your app (repo root / Output)
repo_root = Path(__file__).resolve().parent
OUTPUT_DIR = (repo_root / "Output")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

SHEET_NAME = "MergeData"
TABLE_NAME = "MergeDataTable"

# Header aliases (same style as exporter)
ALIASES = {
    "Name": ["Name", "RecipientFullName", "Defendant Name", "Defendant_Name"],
    "Address_1": ["Address_1", "Address 1", "Address1", "Street", "Street Address"],
    "City": ["City", "City/Town"],
    "State": ["State", "ST", "State Abbrev"],
    "Zip": ["Zip", "ZIP", "Zip Code", "ZipCode", "Postal", "Postal Code"],
    "Case_Number": ["Case_Number", "Case Number", "CaseNumber", "Case No", "CaseNo"],
    "Race": ["Race", "Race/Ethnicity", "Ethnicity", "RACE"],
}

CANON = ["Name","Address_1","City","State","Zip","Case_Number","Race"]

def _canonicalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    colmap = {}
    lower = {str(c).lower(): c for c in df.columns}
    for canon, alist in ALIASES.items():
        for a in alist:
            if a.lower() in lower:
                colmap[lower[a.lower()]] = canon
                break
    df = df.rename(columns=colmap)
    for c in CANON:
        if c not in df.columns:
            df[c] = ""
    return df[CANON]

def _read_best_sheet(path: Path) -> pd.DataFrame:
    xls = pd.ExcelFile(path)
    # try first sheet fast
    try:
        df0 = pd.read_excel(path, sheet_name=xls.sheet_names[0])
        return _canonicalize_columns(df0)
    except Exception:
        pass
    # scan all sheets
    for s in xls.sheet_names:
        try:
            df = pd.read_excel(path, sheet_name=s)
            return _canonicalize_columns(df)
        except Exception:
            continue
    raise ValueError(f"No recognizable headers in {path.name}")

def _fix_zip(s):
    if pd.isna(s):
        return ""
    s = str(s).strip()
    if s.endswith(".0"):
        s = s[:-2]
    if s.isdigit() and len(s) <= 5:
        s = s.zfill(5)
    return s

def pick_excel(title="Pick the merge source Excel"):
    root = tk.Tk(); root.withdraw()
    p = filedialog.askopenfilename(
        title=title, filetypes=[("Excel files","*.xlsx *.xlsm *.xls")]
    )
    root.update_idletasks(); root.destroy()
    return Path(p) if p else None

def pick_template(title="Pick your Word letter template (.docx/.docm)"):
    root = tk.Tk(); root.withdraw()
    p = filedialog.askopenfilename(
        title=title, filetypes=[("Word templates","*.docx *.docm")]
    )
    root.update_idletasks(); root.destroy()
    return Path(p) if p else None

def open_with_word(doc_path: Path):
    """
    On macOS: open the template in Word; you then attach the filtered Excel (Use Existing List…).
    """
    try:
        import subprocess, shlex
        subprocess.run(["open", str(doc_path)], check=False)
    except Exception:
        pass

def save_filtered(df: pd.DataFrame, suffix: str) -> Path:
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    out = df.copy()
    out["Zip"] = out["Zip"].apply(_fix_zip)

    stamp = dt.datetime.now().strftime("%Y-%m-%d")
    outfile = OUTPUT_DIR / f"English_MergeData_{stamp}{suffix}.xlsx"
    with pd.ExcelWriter(outfile, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name=SHEET_NAME)
        # add a Table only if rows >= 1
        ws = writer.book[SHEET_NAME]
        nrows, ncols = out.shape
        if nrows >= 1:
            ref = f"A1:{get_column_letter(ncols)}{nrows+1}"
            table = Table(displayName=TABLE_NAME, ref=ref)
            style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            table.tableStyleInfo = style
            ws.add_table(table)
            # Zip as text
            for cell in ws["E"][1:]:
                cell.number_format = "@"
    return outfile

def main():
    src = pick_excel("Pick the merge source Excel (from Output or your own)")
    if not src:
        print("Cancelled."); return 1
    try:
        df = _read_best_sheet(src)
    except Exception as e:
        tk.Tk().withdraw()
        messagebox.showerror("Error", f"Could not read {src.name}:\n{e}")
        return 2

    # Filter: Race != 'Hispanic' (case-insensitive, trims)
    def is_hispanic(x):
        if pd.isna(x): return False
        return str(x).strip().lower() == "hispanic"

    english_df = df[~df["Race"].apply(is_hispanic)].copy()

    # Save filtered file
    outfile = save_filtered(english_df, suffix="_RaceNotHispanic")

    # Show result + ask for template to open
    try:
        root = tk.Tk(); root.withdraw()
        messagebox.showinfo("Filter complete",
                            f"Saved filtered Excel for English letters:\n{outfile}\n\n"
                            f"Rows: {len(english_df)}\n\n"
                            f"Next you'll pick your Word template so you can Finish & Merge.")
    except Exception:
        pass

    tmpl = pick_template()
    if tmpl:
        open_with_word(tmpl)
        print(f"Opened: {tmpl}")
    else:
        print(f"Saved: {outfile}\n(No template selected to open.)")

    return 0

if __name__ == "__main__":
    sys.exit(main())
