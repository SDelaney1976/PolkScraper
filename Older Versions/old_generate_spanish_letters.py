# src/exports/generate_spanish_letters.py
import os
import sys
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import pandas as pd
from datetime import date

# openpyxl for adding an Excel Table (so Word mail-merge is clean)
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


def _pick_excel(title="Select source Excel (the merged or raw file)"):
    root = tk.Tk(); root.withdraw()
    path = filedialog.askopenfilename(
        title=title,
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"), ("All files", "*.*")]
    )
    root.update(); root.destroy()
    return Path(path) if path else None


def _ask_date(default_str):
    root = tk.Tk(); root.withdraw()
    s = simpledialog.askstring(
        "Capture Date",
        "Enter the capture date (YYYY-MM-DD) for the output filename:",
        initialvalue=default_str,
        parent=None
    )
    root.update(); root.destroy()
    return s.strip() if s else None


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _filter_hispanic(df: pd.DataFrame) -> pd.DataFrame:
    if "Race" not in df.columns:
        raise ValueError("The selected Excel does not contain a 'Race' column.")
    return df[df["Race"].astype(str).str.strip().str.lower() == "hispanic"]


def _write_with_table(df: pd.DataFrame, out_path: Path,
                      sheet_name="MergeData", table_name="MergeDataTable"):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    # 1) Write with pandas
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # 2) Add an Excel Table across the data
    wb = load_workbook(out_path)
    ws = wb[sheet_name]

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 2:  # only header or empty
        wb.save(out_path)
        return

    end_col = get_column_letter(max_col)
    ref = f"A1:{end_col}{max_row}"

    # ensure unique table name if file is reused
    tn = table_name
    existing = {t.displayName for t in ws._tables}
    i = 2
    while tn in existing:
        tn = f"{table_name}_{i}"
        i += 1

    table = Table(displayName=tn, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)
    wb.save(out_path)


def _open_word_template_picker():
    root = tk.Tk(); root.withdraw()
    path = filedialog.askopenfilename(
        title="Open the Spanish Word letter template (.docx/.docm)",
        filetypes=[("Word documents", "*.docx *.docm"), ("All files", "*.*")]
    )
    root.update(); root.destroy()
    return Path(path) if path else None


def main():
    # 1) Pick the source Excel
    src = _pick_excel("Select the Excel to pull Hispanic rows from")
    if not src:
        print("Canceled: no Excel selected.")
        return

    # 2) Ask for date for the output filename (default = today)
    default_date = date.today().strftime("%Y-%m-%d")
    cap_date = _ask_date(default_date)
    if not cap_date:
        print("Canceled: no capture date provided.")
        return

    # 3) Read, normalize, filter rows
    try:
        df = pd.read_excel(src, dtype=str)
        df = _normalize_columns(df)
        df_out = _filter_hispanic(df)
    except Exception as e:
        messagebox.showerror("Read/Filter Error", f"Could not process Excel:\n{e}")
        return

    if df_out.empty:
        messagebox.showinfo("No Rows", "No Hispanic rows found in the selected file.")
        return

    # 4) Build the Output folder path (match English behavior)
    exports_dir = Path(__file__).resolve().parent  # .../src/exports
    FROZEN = bool(getattr(sys, "frozen", False))

    if FROZEN:
        # When frozen, you generally can't write inside the app bundle.
        # Use the same repo-level Output that app.py uses.
        repo_root = Path(sys.executable).resolve().parents[4]
        OUTPUT_DIR = repo_root / "Output"
    else:
        # In dev, write alongside this module: .../src/exports/Output
        OUTPUT_DIR = exports_dir / "Output"

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / f"Spanish_MergeData_{cap_date}.xlsx"

    # 5) Write the filtered data with an Excel Table
    try:
        _write_with_table(df_out, out_path, sheet_name="MergeData", table_name="MergeDataTable")
    except Exception as e:
        messagebox.showerror("Write Error", f"Failed writing output Excel:\n{e}")
        return

    # 6) Prompt to open the Word template (Spanish)
    messagebox.showinfo("Export complete", f"Saved:\n{out_path}\n\nNext: pick your Spanish Word template.")
    tpl = _open_word_template_picker()
    if tpl:
        try:
            os.system(f"open '{tpl.as_posix()}'")  # macOS: open in Word
        except Exception:
            pass  # Excel is ready regardless
    else:
        print("Template open canceled.")


if __name__ == "__main__":
    main()
