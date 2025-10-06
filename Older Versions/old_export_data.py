#!/usr/bin/env python3
"""
export_data.py  (Race support + robust date parsing + zero-row safe)
-------------------------------------------------------------------
Creates a merge-ready Excel for Word letters by combining up to 3 source Excel files,
filtering rows to a chosen Capture Date, and normalizing columns to:
  Name, Address_1, City, State, Zip, Case_Number, Race

Output: <project>/Output/MergeData_YYYY-MM-DD.xlsx (sheet "MergeData"; Excel Table when rows >= 1).

Improvements:
- Accepts datetime strings like "2025-08-15 00:00:00" (time ignored).
- If zero rows match, writes headers only and SKIPS Excel Table (prevents 'repaired content').
- Shows a summary count per source file and total after export.
- Auto-detects the correct sheet in each workbook by scanning for recognizable headers.
"""

import sys
import datetime as dt
from pathlib import Path
from typing import List, Optional, Tuple

import pandas as pd

# GUI pickers (Tkinter is built-in)
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox

# ------- SETTINGS -------
# Use the same Output folder convention as your app.py (repo root / Output)
repo_root = Path(__file__).resolve().parent.parent
OUTPUT_DIR = repo_root / "Output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

OUTPUT_STEM = "MergeData"
SHEET_NAME = "MergeData"
TABLE_NAME = "MergeDataTable"

# Accepted header aliases (case-insensitive)
ALIASES = {
    "Name": ["Name", "RecipientFullName", "Defendant Name", "Defendant_Name"],
    "Address_1": ["Address_1", "Address 1", "Address1", "Street", "Street Address"],
    "City": ["City", "City/Town"],
    "State": ["State", "ST", "State Abbrev"],
    "Zip": ["Zip", "ZIP", "Zip Code", "ZipCode", "Postal", "Postal Code"],
    "Case_Number": ["Case_Number", "Case Number", "CaseNumber", "Case No", "CaseNo"],
    "Capture Date": ["Capture Date", "Capture_Date", "CaptureDate", "Captured Date", "Capture Dt"],
    "Race": ["Race", "Race/Ethnicity", "Ethnicity", "RACE"],
}

CANON = ["Name", "Address_1", "City", "State", "Zip", "Case_Number", "Race", "Capture Date"]

def _canonicalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Map alias columns to canonical names; keep only canonical fields present."""
    colmap = {}
    lower_cols = {str(c).lower(): c for c in df.columns}
    for canon, alist in ALIASES.items():
        found = None
        for a in alist:
            c = lower_cols.get(a.lower())
            if c is not None:
                found = c
                break
        if found:
            colmap[found] = canon
    if "Capture Date" not in colmap.values():
        raise ValueError("Could not find a 'Capture Date' column in this sheet.")
    df = df.rename(columns=colmap)
    keep = [c for c in CANON if c in df.columns]
    return df[keep]

def _read_best_sheet(path: Path) -> pd.DataFrame:
    """
    Try first sheet; if it lacks recognizable headers, scan all sheets and pick the first
    that canonicalizes successfully.
    """
    xls = pd.ExcelFile(path)
    # Try the first sheet
    try:
        df0 = pd.read_excel(path, sheet_name=xls.sheet_names[0])
        try:
            return _canonicalize_columns(df0)
        except Exception:
            pass
    except Exception:
        pass

    # Scan all sheets for one that canonicalizes
    for s in xls.sheet_names:
        try:
            df = pd.read_excel(path, sheet_name=s)
            return _canonicalize_columns(df)
        except Exception:
            continue

    raise ValueError(f"No sheet with recognizable headers found in {path.name}")

def _to_date(x):
    """
    Robustly coerce values like '2025-08-15 00:00:00', Excel serials, python dates, etc. to date().
    """
    if pd.isna(x):
        return pd.NaT
    if isinstance(x, (dt.date, dt.datetime, pd.Timestamp)):
        try:
            return pd.to_datetime(x).date()
        except Exception:
            return pd.NaT
    s = str(x).strip()
    # Drop time portion if present (handles ' ' or 'T')
    for sep in (" ", "T"):
        if sep in s:
            s = s.split(sep, 1)[0]
            break
    try:
        return pd.to_datetime(s, errors="coerce").date()
    except Exception:
        return pd.NaT

def _fix_zip(z):
    if pd.isna(z):
        return ""
    s = str(z).strip()
    if s.endswith(".0"):
        s = s[:-2]
    if s.isdigit() and len(s) <= 5:
        s = s.zfill(5)
    return s

def export_merge_data(input_files: List[Path], merge_date: Optional[dt.date] = None) -> Path:
    """Core export function (callable from app.py)."""
    merge_date = merge_date or dt.date.today()
    frames = []
    per_file_counts: List[Tuple[str, int]] = []

    for f in input_files:
        df = _read_best_sheet(Path(f))
        # Normalize dates and filter
        df["Capture Date"] = df["Capture Date"].apply(_to_date)
        df = df[df["Capture Date"] == merge_date]

        # Ensure required columns exist
        for col in ["Name", "Address_1", "City", "State", "Zip", "Case_Number", "Race"]:
            if col not in df.columns:
                df[col] = ""

        sel = df[["Name", "Address_1", "City", "State", "Zip", "Case_Number", "Race"]].copy()
        sel["Zip"] = sel["Zip"].apply(_fix_zip)

        frames.append(sel)
        per_file_counts.append((Path(f).name, len(sel)))

    if not frames:
        raise RuntimeError("No input files provided.")

    out = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(
        columns=["Name", "Address_1", "City", "State", "Zip", "Case_Number", "Race"]
    )

    # Write to Excel; only create a Table if we have at least 1 data row
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    outfile = OUTPUT_DIR / f"{OUTPUT_STEM}_{merge_date.isoformat()}.xlsx"

    with pd.ExcelWriter(outfile, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name=SHEET_NAME)
        ws = writer.book[SHEET_NAME]
        nrows, ncols = out.shape
        if nrows >= 1:
            ref = f"A1:{get_column_letter(ncols)}{nrows + 1}"
            table = Table(displayName=TABLE_NAME, ref=ref)
            style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            table.tableStyleInfo = style
            ws.add_table(table)
            # Force Zip column as text for data rows
            for cell in ws["E"][1:]:  # E column is Zip
                cell.number_format = "@"
        # else: headers-only, skip table to avoid "repaired content" message

    # Show a summary dialog
    total = sum(c for _, c in per_file_counts)
    try:
        root = tk.Tk(); root.withdraw()
        lines = [f"{name}: {cnt} rows" for name, cnt in per_file_counts]
        lines.append(f"\nTOTAL: {total} rows")
        messagebox.showinfo("Export complete", f"Saved:\n{outfile}\n\n" + "\n".join(lines))
    except Exception:
        pass

    return outfile

# --------------- GUI helpers ---------------
def _pick_files_gui(title="Pick 1–3 Excel files"):
    root = tk.Tk()
    root.withdraw()
    paths = filedialog.askopenfilenames(
        title=title,
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")]
    )
    root.update_idletasks()
    root.destroy()
    return [Path(p) for p in paths]

def _ask_date_gui(default=None):
    root = tk.Tk()
    root.withdraw()
    default = default or dt.date.today().isoformat()
    s = simpledialog.askstring("Capture Date", "Enter date (YYYY-MM-DD). Leave blank to cancel.", initialvalue=default)
    root.update_idletasks()
    root.destroy()
    if not s:
        return None
    try:
        return dt.date.fromisoformat(s.strip())
    except Exception:
        tk.Tk().withdraw()
        messagebox.showerror("Invalid date", "Use format YYYY-MM-DD, e.g., 2025-08-16")
        return _ask_date_gui(default)

def export_today_merge_data_with_gui() -> Path:
    """GUI flow for your button: file picker + date prompt + export; returns output path."""
    files = _pick_files_gui("Pick your 1–3 source Excel files")
    if not files:
        raise RuntimeError("No files selected.")
    if len(files) > 3:
        files = files[:3]
    date_choice = _ask_date_gui(dt.date.today().isoformat())
    if date_choice is None:
        raise RuntimeError("Cancelled by user.")
    out = export_merge_data(files, date_choice)
    return out

# --------------- CLI ---------------
def main(argv=None):
    argv = argv or sys.argv[1:]
    merge_date = None
    # allow --date YYYY-MM-DD anywhere
    if "--date" in argv:
        i = argv.index("--date")
        try:
            merge_date = dt.date.fromisoformat(argv[i+1])
            del argv[i:i+2]
        except Exception:
            print("Invalid --date value. Use YYYY-MM-DD.", file=sys.stderr)
            return 2

    files = [Path(a) for a in argv if a.lower().endswith((".xlsx",".xlsm",".xls"))]

    if not files:
        try:
            out = export_today_merge_data_with_gui()
            print(f"Exported: {out}")
            return 0
        except Exception as e:
            print(f"Error: {e}", file=sys.stderr)
            return 1
    else:
        try:
            out = export_merge_data(files, merge_date)
            print(f"Exported: {out}")
            return 0
        except Exception as e:
            print(f"Error: {e}", file=sys.stderr)
            return 1

if __name__ == "__main__":
    sys.exit(main())
